import requests
import json
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import pandas as pd
import os
from PyWrike.gateways import OAuth2Gateway1
from PyWrike import (
    validate_token,
    authenticate_with_oauth2,
    get_all_spaces,
    get_space_id_from_name,
    get_all_folders,
    get_titles_hierarchy,
    get_custom_statuses,
    create_custom_status_mapping,
    get_custom_fields,
    create_custom_field_mapping,
    get_tasks_for_folder,
    get_tasks_details,
    clean_html,
    get_user_details
)

def wrike_export_main():
    user_cache = {}
    # Load configuration and delete task details from Excel file
    excel_file = input("Enter the path to the Excel file: ")
    if not os.path.isfile(excel_file):
        print("File does not exist. Please check the path.")
        exit()

    try:
        config_df = pd.read_excel(excel_file, sheet_name="Config", header=1)
    except Exception as e:
        print(f"Failed to read Excel file: {e}")
        exit()

    # Extract information from the Config sheet 
    access_token = config_df.at[0, "Token"]
    space_name = config_df.at[0, "Space to extract data from"]

        # Validate the token
    if not validate_token(access_token):
        # If the token is invalid, authenticate using OAuth2 and update the access_token
        wrike = OAuth2Gateway1(excel_filepath=excel_file)
        access_token = wrike._create_auth_info()  # Perform OAuth2 authentication
        print(f"New access token obtained: {access_token}")

    # Fetch all spaces and get the space ID for the given space name
    spaces_response = get_all_spaces(access_token)
    space_id = get_space_id_from_name(space_name, spaces_response)

    if not space_id:
        print(f"Space with name '{space_name}' not found.")
        exit()

    # Extract titles hierarchy for all top-level folders
    print("Extracting titles hierarchy for all top-level folders...")
    folders_response = get_all_folders(space_id, access_token)
    all_paths = []
    processed_folder_ids = set()
    for folder in folders_response["data"]:
        if "scope" in folder and folder["scope"] == "WsFolder":
            paths = get_titles_hierarchy(folder["id"], folders_response["data"])
            all_paths.extend(paths)
    # Fetch custom statuses and create a mapping
    workflows_response = get_custom_statuses(access_token)
    custom_status_mapping = create_custom_status_mapping(workflows_response)

    # Fetch custom fields and create a mapping
    custom_fields_response = get_custom_fields(access_token)
    custom_field_mapping = create_custom_field_mapping(custom_fields_response)

    # Initialize a set to store all unique custom field names
    custom_field_names = set()

    # Extract tasks and subtasks for each folder
    print("Extracting tasks and subtasks for each folder...")
    all_folders_tasks = []
    task_counter = 1  # Counter for task keys
    for folder in all_paths:
        folder_id = folder["id"]
        if folder_id in processed_folder_ids:
            print(f"Skipping already processed folder {folder_id}.")
            continue
        processed_folder_ids.add(folder_id)
        print(f"Processing folder {folder_id} - {folder['path']}...")
        try:
            tasks = get_tasks_for_folder(folder_id, access_token)
            folder_tasks = {
            "folder_id": folder_id,
            "folder_path": folder["path"].replace(f"/{space_name}", "", 1) if folder["path"].startswith(f"/{space_name}/") else folder["path"].replace(f"{space_name}", ""),
            "tasks": []
    }

            print(f"Found {len(tasks)} tasks in folder {folder_id}.")
            for task in tasks:
                task_details = get_tasks_details(task["id"], access_token, custom_status_mapping, custom_field_mapping)
                dates = task_details.get("dates", {})
                start_date = dates.get("start", "")
                due_date = dates.get("due", "")
                duration = dates.get("duration", "")
                efforts = task_details.get("effortAllocation", {})
                effort = efforts.get("totalEffort", "")
                # Clean the HTML content for description
                description_html = task.get("description", "")
                description_cleaned = clean_html(description_html)

                # Fetch emails for responsible IDs
                responsible_emails = []
                for user_id in task_details.get("responsibleIds", []):
                    try:
                        responsible_emails.append(get_user_details(user_id, access_token, user_cache))
                    except requests.exceptions.HTTPError as e:
                        print(f"Error fetching user details for {user_id}: {e}")
                        if e.response.status_code == 429:
                            print("Rate limit exceeded. Sleeping for 60 seconds...")
                            time.sleep(60)
                        responsible_emails.append("Unknown")
                    except Exception as e:
                        print(f"Unexpected error: {e}")
                        responsible_emails.append("Unknown")
                responsible_emails_str = ", ".join(responsible_emails)

                
                # Update the set of custom field names with the current task's custom fields
                for field_id, value in task_details.get("customFields", {}).items():
                    custom_field_names.add(custom_field_mapping.get(field_id, field_id))

                task_key = f"T{task_counter}"
                task_counter += 1
                folder_tasks["tasks"].append({
                    "task_key": task_key,
                    "task_id": task["id"],
                    "task_title": task["title"],
                    "task_description": description_cleaned,
                    "task_responsibleEmails": responsible_emails_str,
                    "status": task_details.get("status", ""),
                    "priority": task_details.get("importance", ""),
                    "custom_status": task_details.get("customStatus", ""),
                    "custom_fields": task_details.get("customFields", {}), 
                    "start_date": start_date,
                    "due_date": due_date,
                    "duration": duration,
                    "effort": effort,
                    "time_spent": task_details.get("timeSpent", ""),
                    "subtasks": []
                })
                subtask_counter = 1
                if "subTaskIds" in task_details:
                    for subtask_id in task_details["subTaskIds"]:
                        subtask_details = get_tasks_details(subtask_id, access_token, custom_status_mapping, custom_field_mapping)
                        subtask_dates = subtask_details.get("dates", {})
                        subtask_start_date = subtask_dates.get("start", "")
                        subtask_due_date = subtask_dates.get("due", "")
                        subtask_duration = subtask_dates.get("duration", "")
                        subtask_efforts = subtask_details.get("effortAllocation", {})
                        subtask_effort = subtask_efforts.get("totalEffort", "")
                        subtask_html = subtask_details.get("description", "")
                        subtask_description_cleaned = clean_html(subtask_html)

                        # Fetch emails for responsible IDs
                        subtask_responsible_emails = []
                        for user_id in subtask_details.get("responsibleIds", []):
                            try:
                                subtask_responsible_emails.append(get_user_details(user_id, access_token, user_cache))
                            except requests.exceptions.HTTPError as e:
                                print(f"Error fetching user details for {user_id}: {e}")
                                if e.response.status_code == 429:
                                    print("Rate limit exceeded. Sleeping for 60 seconds...")
                                    time.sleep(60)
                                subtask_responsible_emails.append("Unknown")
                            except Exception as e:
                                print(f"Unexpected error: {e}")
                                subtask_responsible_emails.append("Unknown")
                        subtask_responsible_emails_str = ", ".join(subtask_responsible_emails)

                        for field_id, value in subtask_details.get("customFields", {}).items():
                            custom_field_names.add(custom_field_mapping.get(field_id, field_id))
                        
                        
                        folder_tasks["tasks"][-1]["subtasks"].append({
                            "subtask_key": f"{task_key}.{subtask_counter}",
                            "subtask_id": subtask_details["id"],
                            "subtask_title": subtask_details["title"],
                            "subtask_description": subtask_description_cleaned,
                            "subtask_responsibleEmails": subtask_responsible_emails_str,
                            "status": subtask_details.get("status", ""),
                            "priority": subtask_details.get("importance", ""),
                            "custom_status": subtask_details.get("customStatus", ""),
                            "subtask_custom_fields": subtask_details.get("customFields", {}),
                            "start_date": subtask_start_date,
                            "due_date": subtask_due_date,
                            "duration": subtask_duration,
                            "effort": subtask_effort,
                            "time_spent": subtask_details.get("timeSpent", "")
                        })
                        subtask_counter += 1
                print(f"Task {task['id']} ({task_details['title']}) with {len(folder_tasks['tasks'][-1]['subtasks'])} subtasks processed.")
            all_folders_tasks.append(folder_tasks)
        except requests.exceptions.HTTPError as e:
            print(f"Error fetching tasks for folder {folder_id}: {e}")
            if e.response.status_code == 429:
                print("Rate limit exceeded. Sleeping for 60 seconds...")
                time.sleep(60)
        except Exception as e:
            print(f"Unexpected error: {e}")

    # Initialize the workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks and Subtasks"

    # Define base headers
    base_headers = ["Key", "Original space name", "Folder", "Parent Task", "Task Title", "Status", "Priority", "Assigned To", "Custom Status", "Start Date", "Duration", "Effort", "Time Spent", "End Date", "Description"]

    # Collect all unique custom field names
    custom_field_names = set()

    # Iterate over folders and tasks
    for folder in folders_response["data"]:
        if "scope" in folder and folder["scope"] == "WsFolder":
            # Get tasks for the current folder
            tasks = get_tasks_for_folder(folder["id"], access_token)
            
            for task in tasks:
                # Get main task details and update custom field names
                task_details = get_tasks_details(task["id"], access_token, custom_status_mapping, custom_field_mapping)
                custom_field_names.update(task_details["customFields"].keys())
                
                # Check if the task has subtasks
                if "subTaskIds" in task:
                    for subtask_id in task["subTaskIds"]:
                        # Get subtask details and update custom field names
                        subtask_details = get_tasks_details(subtask_id, access_token, custom_status_mapping, custom_field_mapping)
                        custom_field_names.update(subtask_details["customFields"].keys())

    # Combine base headers with dynamic custom field headers (name and type)
    headers = base_headers + [custom_field_mapping.get(field_id, field_id) for field_id in custom_field_names]
    ws.append(headers)

    # Populate workbook with tasks and subtasks
    for folder_tasks in all_folders_tasks:
        folder_path = folder_tasks["folder_path"]
        for task in folder_tasks["tasks"]:
            
            task_data = [
                task["task_key"],
                space_name,
                folder_path,
                "",
                task["task_title"],
                task["status"],
                task["priority"],
                task["task_responsibleEmails"],
                task["custom_status"],
            
                task["start_date"],
                task["duration"],
                task["effort"],
                task["time_spent"],
                task["due_date"],
                task["task_description"]
            ]
            
            for field in custom_field_names:
                task_data.append(task["custom_fields"].get(field, ""))  # Use task's custom_fields here
        
            ws.append(task_data)

            for subtask in task["subtasks"]:
            
                subtask_data = [
                    subtask["subtask_key"],
                    space_name,
                    folder_path,
                    task["task_title"],
                    subtask["subtask_title"],
                    subtask["status"],
                    subtask["priority"],
                    subtask["subtask_responsibleEmails"],
                    subtask["custom_status"],
                    
                    subtask["start_date"],
                    subtask["duration"],
                    subtask["effort"],
                    subtask["time_spent"],
                    subtask["due_date"],
                    subtask["subtask_description"]
                ]
                
                # Add custom field values for subtasks
                for field in custom_field_names:
                    subtask_data.append(subtask["subtask_custom_fields"].get(field, ""))  # Use subtask's custom_fields here

                ws.append(subtask_data)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to a file
    output_filename = f"export_{space_name.replace(' ', '_')}.xlsx"
    wb.save(output_filename)
    print(f"Workbook '{output_filename}'")

if __name__ == "__wrike_export_main__":
    wrike_export_main()