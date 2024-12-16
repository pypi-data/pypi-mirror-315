#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
    pd 的 `to_excel()` 函数允许将 df 对象导出到一个 Excel 文件中。

    这个功能是 pd 库中的一部分，使用起来非常方便，
    尤其是在处理数据分析和数据科学任务时。

    以下是一些 `to_excel()` 函数的基本用法和参数：
        - 基本用法:
              `df.to_excel("output.xlsx")`
          这会将 df 保存到名为 "output.xlsx" 的 Excel 文件中。

        - 指定工作表(sheet)名称:
              `df.to_excel("output.xlsx", sheet_name="Sheet1")`
          通过 `sheet_name` 参数可以指定工作表的名称，默认为 "Sheet1"。

        - 设置没有索引:
              `df.to_excel("output.xlsx", index=False)`
          如果不想将 df 的索引保存到 Excel 文件中，
          可以使用 `index=False` 参数。

        - 设置列标题:
              `df.to_excel("output.xlsx", header=True)`
          默认情况下 `header=True`，
          它会将列名作为 Excel 文件中的标题行。
          如果不需要，可以设置为 False。

    pd 的 `to_excel()` 函数是 df 对象的一个非常有用的方法，
    它允许将 df 数据直接写入到 Excel 文件中。

    这个功能是通过依赖 `openpyxl` 或者 `xlsxwriter` 等库实现的，
    为 py 中数据分析提供了极大的便利。

    以下是 `to_excel()` 函数的一些关键特性和参数：
        - `excel_writer`:
          指定输出的 Excel 文件的路径或者现有的 ExcelWriter 对象
          （例如，使用 `pd.ExcelWriter()` 创建）。
          当需要将多个 df 写入同一个文件的不同工作表时，
          可以使用 ExcelWriter。

        - `sheet_name`: 字符串类型，默认为 "Sheet1"。
          指定要写入数据的工作表名称。

        - `na_rep`: 字符串类型，默认为空字符串。
          在写入 Excel 时用来替换缺失值（NaN）。

        - `float_format`: 字符串类型，默认为 None。控制浮点数的格式。

        - `columns`: 序列类型，默认为 None。如果指定，则只输出这些列。

        - `header`: 布尔值或字符串列表，默认为 True。
          如果传递了字符串列表，则假设它们是列名；
          如果 `header` 是 `False`，则在输出文件中省略列名。

        - `index`: 布尔值，默认为 True。决定是否写入索引（行标签）。
          如果设置为 False，则不会在 Excel 文件中显示 df 索引。

        - `index_label`: 字符串或序列，默认为 None。
          如果指定并且 index 为 True，则使用这个参数作为索引列的标签。

        - `startrow` 和 `startcol`:
          分别指定开始写数据时起始行和起始列（从 0 开始计数）。

        - `engine`: 字符串类型，指定用于写文件的引擎，
          比如 "openpyxl" 或 "xlsxwriter" 等。

        - `merge_cells`: 布尔值，默认为 True。
          决定是否合并属于 MultiIndex index 或 columns
          （即具有多级索引）中相同类别下各单元格。

        - `encoding`: 已废弃参数，在新版本 pd 中不再使用，
          因 py 默认已经是 utf-8 编码。

        - `infs_rep`：字符串类型，默认为空字符串。
          定义当数据框中存在无穷大值时将其替换成该参数所表示的字符。

        - `mode`：只在使用 `ExcelWriter` 对象时可用；
          设置模式，例如追加到现有文件而非覆盖等模式选项
          （例如 "a" 表示追加模式）。

    请注意，在执行 `to_excel()` 之前需要确保
    已安装 `openpyxl` 或者 `xlsxwriter` 库，
    因为 pd 需要这些库来处理 xlsx 文件格式。

    通过使用不同参数可以很灵活地控制输出内容和格式，
    并且可以很容易地集成到自动化数据处理流程中去。

    注意：本脚本中出现的 "py"，如无特殊说明，则指代 "Python"。
         本脚本中出现的 "pd"，如无特殊说明，则指代 "Pandas"。
         本脚本中出现的 "df"，如无特殊说明，则指代 "DataFrame"。
         本脚本中出现的 "np"，如无特殊说明，则指代 "Numpy"。
"""
import sys
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
import xlsxwriter

EXCEL_WRITER_ENGINE = "openpyxl"

# `__all__` 是一个特殊的列表
# 它定义当从模块执行 `from module import *` 时应该导入哪些属性
# 如果定义了 `__all__`，只有在这个列表中的属性才会被导入
# 如果没有定义 `__all__`，那么默认导入模块中不以下划线开头的所有属性
__all__ = [
    "pandas2excel"
]

# 将当前运行的 py 文件所在的上两级目录加入到 py 的【系统路径】中
# 使得在这个【根目录】下的【模块】和【包】可以被当前文件所引用
current_file_path = Path(__file__).absolute()
# 移动到上两级目录以获取【根路径】
root_path = current_file_path.parent.parent
# 将【根路径】作为【系统路径】加入 `sys.path`
sys.path.append(str(root_path))


def __set_cell_styles_openpyxl(worksheet: Worksheet) -> None:
    """
        为 openpyxl 模块中的工作表(worksheet)的所有单元格设置统一的样式。

        本函数将为指定的工作表中的每一个单元格配置以下样式：
            - 边框: 所有边框使用细线条(thin)样式。

            - 对齐: 设置文字水平和垂直居中对齐，以及自动换行。

    :param worksheet: 一个 openpyxl 工作表对象，
    即将应用样式的目标工作表。
    :return: 无返回值。
    """
    from openpyxl.styles import Alignment, Border, Side

    # 创建一个边框样式对象，其中所有边框都设置为细线条样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # 创建一个 Alignment 对象来设置单元格文本格式化属性
    # 水平居中、垂直居中，并启用文字自动换行功能
    align_center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = align_center


def __set_cell_styles_xlsxwriter(
        workbook: xlsxwriter.workbook.Workbook,
        worksheet: xlsxwriter.workbook.Worksheet) -> None:
    """
        为指定的工作簿和工作表设置单元格样式。

        该函数基于 xlsxwriter 库，
        为工作簿中的工作表创建并应用一个统一的单元格格式。

        这个格式包括细线边框、水平垂直居中对齐以及自动换行功能。

    :param workbook: xlsxwriter 的 Workbook 对象，
    代表要操作的 Excel 工作簿。
    :param worksheet: xlsxwriter 的 Worksheet 对象，
    代表要设置样式的 Excel 工作表。
    :return: 无返回值。
    """
    # 创建一个格式对象，配置边框为细线（thin）边框和居中对齐
    cell_format = workbook.add_format(
        {
            # 设置细线边框
            "border": 1,
            # 水平居中
            "align": "center",
            # 垂直居中
            "valign": "vcenter",
            # 自动换行
            "text_wrap": True
        }
    )
    # 设置全工作表范围内所有单元格的默认样式
    worksheet.set_column("A:XFD", None, cell_format)


def __compute_column_widths(df: pd.DataFrame) -> np.ndarray:
    """
         根据 df 中的数据计算推荐的列宽。

        该函数会遍历 df 的每一列，包括表头和数据单元格，
        并计算出最适合每列内容显示的宽度。

        为了更准确地反映所需的宽度，
        这里采用了 UTF-8 编码长度作为宽度的估算值。

    :param df: pd.df - 需要计算列宽的 df 对象。
    :return: np.ndarray -
    包含每列推荐宽度（按 UTF-8 编码长度）的 np 数组。
    """
    widths = np.max(
        [
            df.columns.to_series().apply(
                lambda x: len(x.encode("utf-8"))
            ).values,
            df.astype(str).applymap(
                lambda x: len(x.encode("utf-8"))
            ).agg(max).values
        ],
        axis=0
    )
    return widths


def __set_openpyxl_column_widths(
        widths: np.ndarray,
        writer: pd.ExcelWriter,
        sheet_name: str) -> None:
    """
        根据提供的列宽数组，设置指定 Excel 工作表的列宽。

        使用 openpyxl 库设置 pd ExcelWriter 对象中特定工作表的列宽。

        每个列宽将会增加 2 个单位以获得更好的可读性。

        注意：此函数不会保存工作簿；
        调用者应单独调用 writer.save() 方法来保存更改。

    :param widths: numpy 数组，包含要设置的每列的宽度。
    :param writer: pd ExcelWriter 对象，用于写入 Excel 文件。
    :param sheet_name: 字符串，指定要设置列宽的工作表名称。
    :return: 无返回值。
    """
    from openpyxl.utils import get_column_letter

    for i, width in enumerate(widths, start=1):
        writer.sheets[sheet_name].column_dimensions[
            get_column_letter(i)
        ].width = width + 2


def __set_xlsxwriter_column_widths(
        widths: np.ndarray,
        writer: pd.ExcelWriter,
        sheet_name: str) -> None:
    """
        设置使用 xlsxwriter 引擎的 Excel 工作表中各列的宽度。

        根据提供的列宽数组 `widths` 来调整指定
        工作表 `sheet_name` 中各列的宽度。

        此函数专门为使用 xlsxwriter 引擎的 pd.ExcelWriter 编写，
        不适用于其他引擎。

        注意：实际设置的列宽会在指定宽度基础上增加 2 个单位，
        以便留出一些额外空间。

    :param widths: np 数组，包含希望设置的每一列的宽度。
    :param writer: pd 的 ExcelWriter 实例，
    必须已经使用 xlsxwriter 引擎初始化。
    :param sheet_name: 字符串，表示要调整列宽的工作表名称。
    :return: 无返回值（None）。
    """
    worksheet = writer.sheets.get(sheet_name)
    for i, width in enumerate(widths, start=1):
        worksheet.set_column(i - 1, i - 1, width + 2)


column_widths_setters = {
    "xlsxwriter": __set_xlsxwriter_column_widths,
    "openpyxl": __set_openpyxl_column_widths,
}


def pandas2excel(
        df: pd.DataFrame,
        filename: str,
        sheet_name: str = "Sheet1",
        engine: str = EXCEL_WRITER_ENGINE,
        index: bool = False,
        header: bool = True,
        **kwargs) -> None:
    """
        将 pd df 导出为 Excel 文件，并自动调整列宽。

        此函数会计算 df 每列内容的最大字节长度，
        以此估计并设置 Excel 单元格的列宽，
        确保内容在 Execl 表格中能够优雅地展示而不会出现截断。

    :param df: 需要导出的 pd df 对象。
    :param filename:
    导出的 Excel 文件名，包含路径和文件扩展名。
    :param sheet_name:
    Excel 文件中工作表的名称，默认为 "Sheet1"。
    :param engine: 写入 Excel 文件时使用的引擎，
    如 "openpyxl" 或 "xlsxwriter"。
    :param index: 是否将行索引输出到 Excel 文件，
    默认为 False，不输出行索引。
    :param header: 是否将列名作为头部输出到 Excel 文件，
    默认为 True，输出头部。
    :return: 无返回值。函数执行后，
    输入的 pd.DataFrame 数据会被写入指定路径的 Excel 文件中。
    """
    # 计算数据框各列内容所需的最大列宽
    widths = __compute_column_widths(df)
    # 创建一个用于写入 Execl 的上下文环境
    # 并确保文件正确保存关闭
    with pd.ExcelWriter(
            filename,
            engine=engine
    ) as writer:
        # 将数据框导出至 Execl
        # 指定工作表名称并排除索引
        df.to_excel(
            excel_writer=writer,
            sheet_name=sheet_name,
            index=index,
            header=header,
            # 解包额外参数
            **kwargs
        )
        if engine == "openpyxl":
            worksheet = writer.sheets[sheet_name]
            __set_cell_styles_openpyxl(worksheet=worksheet)
        if engine == "xlsxwriter":
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            __set_cell_styles_xlsxwriter(
                workbook=workbook,
                worksheet=worksheet
            )
        if engine in column_widths_setters:
            set_column_widths_func = \
                column_widths_setters.get(engine)
            set_column_widths_func(
                widths=widths,
                writer=writer,
                sheet_name=sheet_name
            )
        else:
            raise ValueError(
                f"不支持的引擎类型 \"{engine}\"。"
            )
