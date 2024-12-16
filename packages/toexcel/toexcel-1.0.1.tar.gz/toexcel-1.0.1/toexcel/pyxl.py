#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
    openpyxl 是一个 py 库，
    用于读取和写入 Excel `2010 xlsx/xlsm/xltx/xltm` 文件。

    这个库可以以编程方式操作 Excel 文件，而不需要安装 Microsoft Excel。

    主要功能包括：
        - 读取工作簿（Excel 文件）：能够加载现有的工作簿文件进行数据抽取。

        - 写入工作簿：能够创建新的 Excel 文件或修改现有文件，并保存更改。

        - 操作工作表：允许添加、删除或重新排序工作表。

        - 单元格操作：可以读取和写入单元格数据，包括文本、数字、日期和公式。

        - 格式设置：支持对单元格字体、颜色、对齐等样式进行设置。

        - 图表创建：能够插入和编辑图表，来可视化数据。

        - 数据验证和筛选：提供了数据验证和自动过滤功能，
          可以实现复杂的数据管理任务。

        - 合并与拆分单元格：允许合并多个单元格或将已合并的单元格分开。

    openpyxl 是处理 Excel文件的有效工具，
    特别适用于需要自动化处理大量电子表格数据的场景。

    注意：本脚本中出现的 "py"，如无特殊说明，则指代 "Python"。
"""
import sys
from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_WIDTH = 35
# `__all__` 是一个特殊的列表
# 它定义当从模块执行 `from module import *` 时应该导入哪些属性
# 如果定义了 `__all__`，只有在这个列表中的属性才会被导入
# 如果没有定义 `__all__`，那么默认导入模块中不以下划线开头的所有属性
__all__ = [
    "create_excel"
]

# 将当前运行的 py 文件所在的上两级目录加入到 py 的【系统路径】中
# 使得在这个【根目录】下的【模块】和【包】可以被当前文件所引用
current_file_path = Path(__file__).absolute()
# 移动到上两级目录以获取【根路径】
root_path = current_file_path.parent.parent
# 将【根路径】作为【系统路径】加入 `sys.path`
sys.path.append(str(root_path))


def __set_styles(
        ws: Worksheet,
        default_width: int = DEFAULT_WIDTH) -> None:
    """
        为传入的工作表设置单元格样式，
        包括文本的对齐方式、自动换行以及边框样式。

        功能：
            - 将所有单元格的对齐方式设置为水平居中和垂直居中。

            - 为所有单元格启用自动换行。

            - 设置所有单元格的边框为细线样式。

            - 将所有列的宽度设置为 35。

        如果遇到合并单元格，将处理其所在列的第一个单元格以应用相同的样式。

        其他部分合并后自动继承相同样式。

    :param ws: openpyxl.workbook.worksheet.Worksheet 的工作表实例，
    将对其进行样式设置。
    :param default_width: 设置工作表中每一列宽度的默认值，
    默认情况下使用常量 DEFAULT_WIDTH。
    如果调用时提供了其他值，则使用该值覆盖默认宽度。
    宽度单位通常是字符数，具体数值取决于使用的字体和 Excel 版本。
    :return: 无返回值。
    """
    # 创建一个 Alignment 对象来设置单元格文本格式化属性
    # 水平居中、垂直居中，并启用文字自动换行功能
    align_center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )
    # 创建一个边框样式对象，其中所有边框都设置为细线条样式
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # 遍历工作表中的每一列
    # `iter_cols()` 方法返回工作表中所有列的迭代器
    for col in ws.iter_cols():
        # 检查当前列的第一个单元格（`col[0]`）是否为合并单元格的一部分
        # `MergedCell` 是 openpyxl 定义的类型，表示合并单元格
        if isinstance(col[0], MergedCell):
            # 如果第一个单元格是合并单元格
            # 调用 `get_column_letter()` 函数来获取当前处理列的列字母标识
            # 例如 "A", "B", "C"
            column_letter = get_column_letter(col[0].column)
        else:
            # 如果第一个单元格不是合并单元格
            # 则直接获取该非合并单元格的列字母标识属性
            column_letter = col[0].column_letter
        ws.column_dimensions[column_letter].width = default_width
        for cell in col:
            if not isinstance(cell, MergedCell):
                cell.alignment = align_center
                cell.border = thin_border


def __add_rows(
        ws: Worksheet,
        rows: List[List],
        field_names: Optional[List[str]] = None) -> None:
    """
        将字段名和多行数据添加到工作表中。

        该函数接收一个工作表对象、字段名列表和数据行列表，
        首先将字段名追加到工作表的顶部，
        然后依次将每一行数据追加到工作表中。

        此函数无返回值。

    :param ws: Worksheet 对象，
    表示需要添加数据的工作表。
    :param field_names: 字符串列表，
    表示要添加到工作表顶部的字段名称。
    :param rows: 嵌套列表，
    其内部每个子列表代表一行数据。
    :return: 无返回值。
    """
    if field_names:
        ws.append(field_names)
    for row in rows:
        ws.append(row)


def create_excel(
        filename: str,
        rows: List[List],
        field_names: Optional[List[str]] = None,
        default_width: int = DEFAULT_WIDTH,
        sheet_name: str = "Sheet1") -> None:
    """
        创建一个 Excel 文件并根据提供的数据填充工作表。

        该函数首先创建一个新的工作簿，并设置当前活跃的工作表。

        然后，将字段名列表添加到工作表的第一行作为标题行。

        随后，对于提供的每一行数据，该函数将其追加到工作表中。

        同时，为了改善在 Excel 中的可读性和格式化，
        该函数为所有单元格设置文本对齐方式、启用自动换行，
        并应用细边框样式。

        最终，函数将工作簿保存到指定的文件名中。

    :param field_names: 字段名列表，
    这些字段名将被写入 Excel 的第一行。
    :param filename: 要保存的 Excel 文件名称。
    :param rows: 列表形式的数据行集合，
    在 Excel 文件中每个元素都会占据一行。
    :param default_width: 设置工作表中每一列宽度的默认值，
    默认情况下使用常量 DEFAULT_WIDTH。
    如果调用时提供了其他值，则使用该值覆盖默认宽度。
    宽度单位通常是字符数，
    具体数值取决于使用的字体和 Excel 版本。
    :param sheet_name:
    Excel 文件中工作表的名称，默认为 "Sheet1"。
    :return: None
    """
    if not filename.endswith(r".xlsx"):
        raise ValueError(
            "文件名必须以 \".xlsx\" 结尾。"
        )
    if field_names is not None \
            and not isinstance(field_names, list):
        raise ValueError(
            "\"field_names\" 必须是个 \"list\"，如果提供的话。"
        )
    wb = Workbook()
    ws = wb.create_sheet(title=sheet_name)
    __add_rows(
        ws=ws,
        field_names=field_names,
        rows=rows
    )
    __set_styles(
        ws=ws,
        default_width=default_width
    )
    wb.save(filename=filename)
