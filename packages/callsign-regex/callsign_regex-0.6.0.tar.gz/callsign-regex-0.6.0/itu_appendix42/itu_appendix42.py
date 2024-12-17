""" itu_appendix42 """

# Table of International Call Sign Series (Appendix 42 to the RR)

# Based on this page ...
#     https://www.itu.int/en/ITU-R/terrestrial/fmd/Pages/glad.aspx
# Visit the following page ...
#     https://www.itu.int/en/ITU-R/terrestrial/fmd/Pages/call_sign_series.aspx
# Download via the .xlsx button and produce a file like this ...
#     CallSignSeriesRanges-998049b7-c007-4e71-bac6-d2393eaa83ef.xlsx
#     CallSignSeriesRanges-c3ce6efb-d36c-4e44-8fff-083b4aab1c09.xlsx
# The following code looks for the newest file of that name pattern in your Download's directory.
# Under windows that's C:\Users\YourUsername\Downloads\, under Linux or MacOS it's ~/Downloads

import sys
import re
from os import makedirs
from os.path import join, expanduser, getmtime, isdir, exists

from glob import glob
from string import ascii_uppercase, digits

try:
    from importlib.resources import files
except ImportError:
    from importlib_resources import files

from openpyxl import load_workbook

from itu_appendix42.iso3661_mapping import iso3661_mapping
from itu_appendix42.iso3661_mapping_from_itu import iso3661_mapping_from_itu

class ItuAppendix42():
    """ ItuAppendix42 """

    DOWNLOAD_FOLDER = 'Downloads'
    FILENAME_PATTERN = 'CallSignSeriesRanges-*-*-*-*.xlsx'
    PACKAGE_RESOURCES = 'itu_appendix42.resources'

    CACHE_FOLDER = '.cache'
    CACHE_SUBFOLDER = 'itu_appendix42'
    CACHE_FILENAME = 'itu_appendix42.regex'

    _forward = None
    _reverse = None
    _regex = None
    _regex_c = None

    _verbose = False

    def __init__(self, force=False, verbose=False):
        """ __init__ """

        self.__class__._verbose = verbose

        if not force:
            # grab cached regex if we can
            regex, regex_mtime = self.__class__._read_cache_regex()
            if regex:
                self.__class__._regex = regex

        # if no cached regex - read in the worksheet(s) and build it manually
        if not self.__class__._regex:
            self.__class__._read_in_all_worksheets()
            # regex build time
            if not self.__class__._regex:
                self.__class__._build_regex()
            # save a local version away in users cache for next run
            self.__class__._write_cache_regex()

        # regex compile time
        self.__class__._regex_c = re.compile(self.__class__._regex, re.ASCII|re.IGNORECASE)

    def regex(self):
        """ regex """
        return self.__class__._regex

    def regex_c(self):
        """ regex_c """
        return self.__class__._regex_c

    def dump(self):
        """ dump """
        # can only dump if we have worksheets - ignore force flag - TODO optimize this!
        if not self.__class__._forward:
            self.__class__._read_in_all_worksheets()
        results = ''
        for k in sorted(self.__class__._forward):
            callsign = k
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            country = self.__class__._forward[k]
            results += '%-10s : %s\n' % (callsign, country)
        return results

    def rdump(self):
        """ rdump """
        # can only dump if we have worksheets - ignore force flag - TODO optimize this!
        if not self.__class__._forward:
            self.__class__._read_in_all_worksheets()
        # we build a reverse maping as we need it
        if not self.__class__._reverse:
            self.__class__._build_reverse()
        results = ''
        for k in sorted(self.__class__._reverse):
            results += '%-70s : %s\n' % (k, ','.join(self.__class__._reverse[k]))
        return results

    def match(self, line):
        """ match """
        return self.__class__._regex_c.match(line.upper())

    def fullmatch(self, line):
        """ match """
        return self.__class__._regex_c.fullmatch(line.upper())

    def findall(self, line):
        """ findall """
        return [''.join(v) for v in self.__class__._regex_c.findall(line.upper())]

    @classmethod
    def _build_forward(cls, ws):
        """ _build_forward """

        def _optimize_callsign(callsign_series):
            """ _optimize_callsign """
            # ['5XA - 5XZ']
            callsign_low, callsign_high = callsign_series.split(' - ')
            # each is three char's long

            if callsign_low[2] == 'A' and callsign_high[2] == 'Z':
                if callsign_low[0:2] == callsign_high[0:2]:
                    return callsign_low[0:2] + '[A-Z]'
                if callsign_low[1] == 'A' and callsign_high[1] == 'Z' and callsign_low[0:1] == callsign_high[0:1]:
                    return callsign_low[0:1] + '[A-Z][A-Z]'
                if callsign_low[1] == '0' and callsign_high[1] == '9' and callsign_low[0:1] == callsign_high[0:1]:
                    return callsign_low[0:1] + '[0-9][A-Z]'

            # For Egypt, Fiji, etc there could be an A-M & N-Z split on the third letter!
            if callsign_low[2] == 'A' and callsign_high[2] == 'M' and callsign_low[0:2] == callsign_high[0:2]:
                if callsign_low[0:2] == callsign_high[0:2]:
                    return callsign_low[0:2] + '[A-M]'

            if callsign_low[2] == 'N' and callsign_high[2] == 'Z' and callsign_low[0:2] == callsign_high[0:2]:
                if callsign_low[0:2] == callsign_high[0:2]:
                    return callsign_low[0:2] + '[N-Z]'

            # return callsign_series using its orginal values - no optimize available
            return callsign_low + ' - ' + callsign_high

        cls._forward = {}
        for v in list(ws.values)[1:]:
            callsign = v[0]
            country = v[1]
            callsign = _optimize_callsign(callsign)
            if country in iso3661_mapping_from_itu:
                country = iso3661_mapping_from_itu[country]['iso3661'] + '/' + iso3661_mapping_from_itu[country]['iso3661_name']
            cls._forward[callsign] = country

    @classmethod
    def _optimize_duplicates(cls):
        """ _optimize_duplicates """

        def dedup(letter_1, letter_2_begin, letter_2_end, present_country):
            """ dedup """
            letter_2 = letter_2_begin
            while letter_2 <= letter_2_end:
                callsign = letter_1 + letter_2 + '[A-Z]'
                del cls._forward[callsign]
                letter_2 = chr(ord(letter_2) + 1)
            if letter_2_begin == letter_2_end:
                letter_range = letter_2_begin
            else:
                letter_range = '[%s-%s]' % (letter_2_begin, letter_2_end)
            callsign = letter_1 + letter_range + '[A-Z]'
            cls._forward[callsign] = present_country

        # now look for second letter sequences
        for letter_1 in sorted(set([v[0:1] for v in cls._forward])):
            for seq in [digits, ascii_uppercase]:
                present_country = None
                letter_2_begin = None
                letter_2_end = None
                for letter_2 in seq:
                    callsign = letter_1 + letter_2 + '[A-Z]'
                    if callsign not in cls._forward:
                        # quite common - this is a non allocated letter sequence
                        if present_country:
                            dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                        present_country = None
                        letter_2_begin = None
                        letter_2_end = None
                        continue
                    if not present_country:
                        # first find of country
                        present_country = cls._forward[callsign]
                        letter_2_begin = letter_2
                        letter_2_end = letter_2
                        continue
                    if present_country != cls._forward[callsign]:
                        # changed country
                        dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                        present_country = cls._forward[callsign]
                        letter_2_begin = letter_2
                        letter_2_end = letter_2
                        continue
                    # continuing country
                    letter_2_end = letter_2

                if present_country:
                    dedup(letter_1, letter_2_begin, letter_2_end, present_country)
                    present_country = None
                    letter_2_begin = None
                    letter_2_end = None

    @classmethod
    def _build_reverse(cls):
        """_build_reverse """
        cls._reverse = {}
        for k in sorted(cls._forward):
            callsign = k
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            if callsign[-5:] == '[A-Z]':
                callsign = callsign[:-5]
            country = cls._forward[k]
            if country not in cls._reverse:
                cls._reverse[country] = []
            cls._reverse[country].append(callsign)

    @classmethod
    def _build_regex(cls):
        """ _build_regex """

        def expand(s):
            """ expand """
            if len(s) != 3:
                return s
            begin = s[0]
            end = s[2]
            s = ''
            c = begin
            while c <= end:
                s += c
                c = chr(ord(c) + 1)
            return s

        one_letter_alpha = '[' + ''.join(sorted([v[0:1] for v in cls._forward if v[0].isalpha() and v[-10:] == '[A-Z][A-Z]'])) + ']'
        one_letter_numeric = '[' + ''.join(sorted([v[0:1] for v in cls._forward if v[0].isnumeric() and v[-10:] == '[A-Z][A-Z]'])) + ']'

        if len(one_letter_alpha) == 3:
            one_letter_alpha = one_letter_alpha[1]
        if len(one_letter_numeric) == 3:
            one_letter_numeric = one_letter_numeric[1]

        two_letters = []
        twos = sorted([v[0:2] for v in cls._forward if v[-10:] != '[A-Z][A-Z]' and v[-5:] in ['[A-Z]', '[A-M]', '[N-Z]']])
        for letter_1 in sorted(set([v[0:1] for v in twos])):
            step1 = sorted([v[1:-5] for v in cls._forward if v[0] == letter_1 and v[-5:] == '[A-Z]'])
            step2 = [v[0] for v in step1 if len(v) == 1]
            step3 = [expand(v[1:4]) for v in step1 if len(v) != 1]
            step4 = ''.join(sorted(step2 + step3))
            # The following is in a specific order
            # While this could (and should) be better code, there's only some very specific patterns in-use presently
            swaps = [
                ['ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'A-Z'],
                ['ABCDEFGHIJKLMNOPQRSTUVWXY', 'A-Y'],
                ['ABCDEFGHIJKLMNOPQRTUVWXYZ', 'A-RT-Z'],
                ['ABCDEFGHIJKLMOPQRSTUVWXYZ', 'A-MO-Z'],
                ['ABCEFGHIJKLMNOPQRSTUVWXYZ', 'A-CE-Z'],
                ['23456789', '2-9'],
                ['2345678', '2-8'],
                ['234567', '2-7'],
                ['2346789', '2-46-9'],
                ['2356789', '2-35-9'],
            ]
            for swap in swaps:
                step4 = step4.replace(swap[0], swap[1])
            two_letter = letter_1 + '[' + step4 + ']'
            two_letters.append(two_letter)

        # there's further optimization that can be done.

        two_letters_split = [(v[0], v[1:]) for v in two_letters]

        two_letters_sorted = {}
        for v in two_letters:
            a = v[0]
            b = v[1:]
            if b in two_letters_sorted:
                two_letters_sorted[b].append(a)
            else:
                two_letters_sorted[b] = [a]

        two_letters = []
        for k,v in two_letters_sorted.items():
            a = ''.join(v)
            if len(a) > 1:
                a = '[' + a + ']'
            two_letters.append(a + k)

        three_letters = sorted([v for v in cls._forward if v[-10:] != '[A-Z][A-Z]' and v[-5:] not in ['[A-Z]', '[A-M]', '[N-Z]']])

        # split these three patterns because we need to know the ones with numbers in the second position need a different pattern format

        two_letters_has_numeric = []
        two_letters_only_alpha = []
        for element in two_letters:
            a = element.split('[')
            if a[0] == '':
                del a[0]
            b = [v for v in a[1] if v.isnumeric()]
            if len(b) > 0:
                two_letters_has_numeric.append(element)
            else:
                two_letters_only_alpha.append(element)

        # Only found 56789 to optimize - but wrote more generic code anyway
        for ii in range(len(two_letters_has_numeric)):
            swaps = [
                ['56789', '5-9'],
            ]
            for swap in swaps:
                two_letters_has_numeric[ii] = two_letters_has_numeric[ii].replace(swap[0], swap[1])

        for ii in range(len(two_letters_only_alpha)):
            swaps = [
                ['56789', '5-9'],
            ]
            for swap in swaps:
                two_letters_only_alpha[ii] = two_letters_only_alpha[ii].replace(swap[0], swap[1])

        # we combine these three patterns. we add the missing letters. We take care of the numbers carefully

        prefix1_letters = [one_letter_numeric + '[A-Z]{1,2}'] + \
                          [one_letter_alpha + '[A-Z]{0,2}'] + \
                          [v + '[A-Z]{0,1}' for v in two_letters_only_alpha] + \
                          three_letters

        prefix2_letters = [v + '[A-Z]{0,1}' for v in two_letters_has_numeric]

        cls._regex = '(' + \
                        '(' + '|'.join(prefix1_letters) + ')(' + '[0-9][0-9A-Z]{0,3}[A-Z]' + ')' + \
                     '|' + \
                        '(' + '|'.join(prefix2_letters) + ')(' + '[0-9A-Z]{0,3}[A-Z]' + ')' + \
                     ')'

    @classmethod
    def _find_cache_file(cls):
        """ _find_cache_file """
        output_file_folder = join(expanduser('~'), cls.CACHE_FOLDER)
        if not isdir(output_file_folder):
            makedirs(output_file_folder)
        output_file_folder = join(output_file_folder, cls.CACHE_SUBFOLDER)
        if not isdir(output_file_folder):
            makedirs(output_file_folder)
        cache_file = join(output_file_folder, cls.CACHE_FILENAME)
        return cache_file

    @classmethod
    def _write_cache_regex(cls):
        """ _write_cache_regex """

        cache_file = cls._find_cache_file()
        with open(cache_file, 'w', encoding='utf-8') as fd:
            fd.write(cls._regex)
            fd.write('\n')

    @classmethod
    def _read_cache_regex(cls):
        """ _read_cache_regex """

        filename = cls._find_cache_file()
        try:
            regex = open(filename, 'r', encoding='utf-8').read().strip()
            if len(regex) == 0:
                return None, None
        except FileNotFoundError:
            return None, None
        try:
            regex_mtime = getmtime(filename)
        except FileNotFoundError:
            return None, None
        return regex, regex_mtime

    @classmethod
    def _package_folder(cls):
        """ _package_folder """
        return str(files(cls.PACKAGE_RESOURCES)._paths[0])

    @classmethod
    def _download_folder(cls):
        """ _download_folder """
        return str(join(expanduser('~'), cls.DOWNLOAD_FOLDER))

    @classmethod
    def _find_filename(cls):
        """ _find_filename """

        #
        # Change in version 3.10 becuase of the added the root_dir and dir_fd parameters ...
        # filenames = glob(cls.FILENAME_PATTERN, root_dir=cls._package_folder()) + glob(cls.FILENAME_PATTERN, root_dir=dirname)
        # ... but we leave it in the old format so it works on earlier Python versions.
        #

        a = []
        for filename in glob(cls._package_folder() + '/' + cls.FILENAME_PATTERN) + glob(cls._download_folder() + '/' + cls.FILENAME_PATTERN):
            mtime = getmtime(filename)
            a.append((filename, mtime))
        all_filenames = sorted(a, key=lambda item: item[1])
        if cls._verbose:
            for f in all_filenames:
                print('DEBUG: %12d older: %s' % (all_filenames[-1][1] - f[1], f[0]), file=sys.stderr)
        best = all_filenames[-1]
        filename = best[0]
        return filename

    @classmethod
    def _find_worksheet(cls):
        """ _find_worksheet """
        try:
            filename = cls._find_filename()
        except FileNotFoundError as e:
            raise FileNotFoundError(cls_.FILENAME_PATTERN) from None
        wb = load_workbook(filename=filename, data_only=True)
        if 'Exported data' != wb.sheetnames[0]:
            raise FileNotFoundError(filename)
        #ws = wb.active
        ws = wb['Exported data']
        return ws

    @classmethod
    def _read_in_all_worksheets(cls):
        """ _read_in_all_worksheets """
        ws = cls._find_worksheet()
        # forward mapping build first
        cls._build_forward(ws)
        # Further processing reduces this data using regex definition methods
        cls._optimize_duplicates()

