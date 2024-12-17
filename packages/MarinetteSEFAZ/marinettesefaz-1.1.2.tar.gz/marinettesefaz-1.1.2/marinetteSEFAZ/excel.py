from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


def salvarPlanilha(df, caminho, sheet):
    #SALVA A TABELA SEM APAGAR AS OUTRAS
    writer = pd.ExcelWriter(caminho, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.close()

    planilha = load_workbook(caminho)
    tabela = planilha[sheet]

    #FORMULA PARA PREENCHER A COLUNA DE PRAZO
    for linha in range(2,tabela.max_row + 1):
        celulaD = tabela[f"D{linha}"]
        celulaC = tabela[f"C{linha}"]
        if celulaC.value == "-" or celulaC.value == "SEM VALIDADE":
            celulaD.value = "SEM PRAZO"
        elif celulaC.value == "PAGO":
            pass
        elif celulaD.value == "ERRO NO PAGAMENTO":
            pass
        else:
            celulaD.value = f'=C{linha}-TODAY()'

    prazo = f'D2:D{tabela.max_row}'
    #CORES PARA PINTAR AS CELULAS
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFF00", fill_type="solid")
    #REGRAS PARA PINTAR AS COLUANS
    red_rule = CellIsRule(operator='lessThanOrEqual', formula=['15'], stopIfTrue=True, fill=red_fill)
    green_rule = CellIsRule(operator='greaterThanOrEqual', formula=['31'], stopIfTrue=True, fill=green_fill)
    yellow_rule = CellIsRule(operator='between', formula=['15','31'], stopIfTrue=True, fill=yellow_fill)
    error_rule = CellIsRule(operator='equal', formula=['"ERRO NO PAGAMENTO"'], stopIfTrue=True, fill=red_fill)

    if tabela.max_row > 1:
            tabela.conditional_formatting.add(prazo, red_rule)
            tabela.conditional_formatting.add(prazo, yellow_rule)
            tabela.conditional_formatting.add(prazo, error_rule)
            tabela.conditional_formatting.add(prazo, green_rule)


    #ALINHAR TAMANHO DAS CELULAS
    for column in tabela.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        tabela.column_dimensions[column_letter].width = adjusted_width

    planilha.save(caminho)
    planilha.close()
    
    
def copiarPlanilha(caminhoOrigem,caminhoDestino):
    planilha = load_workbook(caminhoOrigem)
    planilha.save(caminhoDestino)  